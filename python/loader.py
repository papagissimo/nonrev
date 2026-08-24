import sqlite3
import openpyxl
from datetime import datetime

DB_PATH = "nonrev.db"
FUZZY_TOLERANCE_MIN = 30


def format_date(year, monthday):
    monthday = int(monthday)
    return f"{int(year):04d}-{monthday // 100:02d}-{monthday % 100:02d}"


def format_hhmm(value):
    # Always normalizes, regardless of whether openpyxl handed us an int,
    # a float (700.0), or a string ("0700" or the corrupted "700.0") -
    # int(float(x)) collapses all of those to the same integer first.
    return f"{int(float(value)):04d}" if value is not None else None


def normalize_flight_number(value):
    # flightNumber is nominal (an identifier, not a quantity) - openpyxl
    # sometimes hands back a real numeric flight number as a Python float
    # (2142.0), which str()'d straight into the db as the literal text
    # "2142.0". Strip that artifact; leave text placeholders like
    # "bogus029" untouched.
    if value is None:
        return None
    s = str(value)
    if s.endswith('.0'):
        try:
            return str(int(float(s)))
        except ValueError:
            return s
    return s


def day_of_week(year, monthday):
    monthday = int(monthday)
    d = datetime(int(year), monthday // 100, monthday % 100)
    return d.strftime("%a")


def hhmm_to_minutes(hhmm_str):
    return int(hhmm_str[:2]) * 60 + int(hhmm_str[2:])


def load_flight_schedule(ws, conn):
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
    skipped = [r for r in rows if not r[1]]
    if skipped:
        print(f"Skipping {len(skipped)} FlightSchedule row(s) with no flightNumber:")
        for r in skipped:
            print("  ", r[:8])
    rows = [r for r in rows if r[1]]
    data = [
        (r[0], normalize_flight_number(r[1]), r[2], r[3], r[4],
         hhmm_to_minutes(format_hhmm(r[5])), r[6], bool(r[7]))
        for r in rows
    ]
    conn.executemany(
        """INSERT INTO flightSchedule
           (carrier, flightNumber, org, dest, dayOfWeek, depTime, aircraftConfig, confirmed, classification)
           VALUES (?,?,?,?,?,?,?,?,NULL)""",
        data,
    )
    # keyed by (org, dest, dayOfWeek) -> list of (depTimeMinutes, carrier, flightNumber)
    # data is already normalized above (depTime is minutes-since-midnight), no further fixing needed.
    by_route_day = {}
    for r in data:
        carrier, flightNumber, org, dest, dow, depTimeMinutes = r[0], r[1], r[2], r[3], r[4], r[5]
        key = (org, dest, dow)
        by_route_day.setdefault(key, []).append(
            (depTimeMinutes, carrier, flightNumber)
        )
    return by_route_day


def find_match(by_route_day, org, dest, dow, dep_time):
    candidates = by_route_day.get((org, dest, dow), [])
    if not candidates:
        return None, None
    target = hhmm_to_minutes(dep_time)
    scored = sorted(candidates, key=lambda c: abs(c[0] - target))
    best = scored[0]
    best_diff = abs(best[0] - target)
    if best_diff > FUZZY_TOLERANCE_MIN:
        return None, None
    if len(scored) > 1:
        second_diff = abs(scored[1][0] - target)
        if second_diff <= FUZZY_TOLERANCE_MIN:
            return None, "ambiguous"  # two candidates both plausibly close
    return (best[1], best[2]), best_diff


def load_aircraft_configs(ws, conn):
    raw_rows = [r[:10] for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
    rows = []
    for r in raw_rows:
        r = list(r)
        for i in range(2, 7):  # d1, first, comfortPlus, main, total
            if r[i] is None:
                r[i] = 0
        rows.append(tuple(r))
    conn.executemany(
        """INSERT INTO aircraftConfigs
           (configKey, aircraft, d1, first, comfortPlus, main, total, status, note, source)
           VALUES (?,?,?,?,?,?,?,?,?,?)""",
        rows,
    )


def load_route_durations(ws, conn):
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
    data = [(r[0], r[1], r[2], bool(r[3])) for r in rows]
    conn.executemany(
        "INSERT INTO routeDurations (org, dest, durationMinutes, confirmed) VALUES (?,?,?,?)",
        data,
    )


def load_observations(ws, conn, by_route_day, carrier_col_index):
    inserted = 0
    unmatched = []
    ambiguous = []
    fuzzy_matches = []
    bad_rows = []
    for row_num, r in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        org = r[1]
        if not org:
            continue
        dest = r[2]
        check_y, check_md, check_hhmm = r[3], r[4], r[5]
        flt_y, flt_md = r[6], r[7]
        dep_raw = r[9]
        hours_before_dep = r[11]
        if not (check_y and check_md and flt_y and flt_md and dep_raw):
            continue

        try:
            check_timestamp = f"{format_date(check_y, check_md)} {format_hhmm(check_hhmm)[:2]}:{format_hhmm(check_hhmm)[2:]}"
            flight_date = format_date(flt_y, flt_md)
            dep_time = format_hhmm(dep_raw)
            dow = day_of_week(flt_y, flt_md)
        except (ValueError, TypeError) as e:
            bad_rows.append((row_num, org, dest, r[3:8], str(e)))
            continue

        match, diff = find_match(by_route_day, org, dest, dow, dep_time)
        if match:
            carrier, flight_number = match
            if diff > 0:
                fuzzy_matches.append((org, dest, dow, dep_time, flight_number, diff))
        else:
            carrier = r[carrier_col_index] or "dl"
            flight_number = None
            if diff == "ambiguous":
                ambiguous.append((org, dest, dow, dep_time, flight_date))
            else:
                unmatched.append((org, dest, dow, dep_time, flight_date))

        blocks = [
            ("avail", r[13], r[14], r[15], r[16]),
            ("soloSelect", r[23], r[24], r[25], r[26]),
            ("pairSelect", r[29], r[30], r[31], r[32]),
        ]
        for reading_type, y, cplus, firstps, d1 in blocks:
            if all(v is None for v in (y, cplus, firstps, d1)):
                continue
            conn.execute(
                """INSERT INTO observations
                   (carrier, flightNumber, org, dest, flightDate, checkTimestamp,
                    hoursBeforeDep, readingType, y, cPlus, firstOrPS, d1)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                (carrier, flight_number, org, dest, flight_date, check_timestamp,
                 hours_before_dep, reading_type, y, cplus, firstps, d1),
            )
            inserted += 1
    return inserted, unmatched, ambiguous, fuzzy_matches, bad_rows


def main(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    conn = sqlite3.connect(DB_PATH)

    by_route_day = load_flight_schedule(wb["FlightSchedule"], conn)
    load_aircraft_configs(wb["AircraftConfigs"], conn)
    load_route_durations(wb["RouteDurations"], conn)

    canbuy = load_observations(wb["canBuy"], conn, by_route_day, carrier_col_index=0)
    archive = load_observations(wb["Archive"], conn, by_route_day, carrier_col_index=33)

    conn.commit()

    canbuy_count, canbuy_unmatched, canbuy_ambig, canbuy_fuzzy, canbuy_bad = canbuy
    archive_count, archive_unmatched, archive_ambig, archive_fuzzy, archive_bad = archive

    total_unmatched = canbuy_unmatched + archive_unmatched
    total_ambiguous = canbuy_ambig + archive_ambig
    total_fuzzy = canbuy_fuzzy + archive_fuzzy
    total_bad = canbuy_bad + archive_bad

    print(f"Loaded {canbuy_count} observations from canBuy, {archive_count} from Archive.")
    print(f"Matched exactly: {canbuy_count + archive_count - len(total_fuzzy) - len(total_unmatched) - len(total_ambiguous)}")
    print(f"Matched within {FUZZY_TOLERANCE_MIN} min (fuzzy): {len(total_fuzzy)}")
    if total_fuzzy:
        print("  Sample fuzzy matches (org, dest, dow, observedDepTime, matchedFlightNumber, diffMinutes):")
        for row in total_fuzzy[:10]:
            print("   ", row)

    print(f"Unmatched (nothing within {FUZZY_TOLERANCE_MIN} min): {len(total_unmatched)}")
    if total_unmatched:
        print("  First 10 (org, dest, dayOfWeek, depTime, flightDate):")
        for row in total_unmatched[:10]:
            print("   ", row)

    print(f"Ambiguous (2+ schedule entries equally plausible): {len(total_ambiguous)}")
    if total_ambiguous:
        print("  First 10 (org, dest, dayOfWeek, depTime, flightDate):")
        for row in total_ambiguous[:10]:
            print("   ", row)

    print(f"Skipped (malformed date/time data): {len(total_bad)}")
    if total_bad:
        for row in total_bad:
            print("  ", row)

    conn.close()


if __name__ == "__main__":
    import sys
    main(sys.argv[1] if len(sys.argv) > 1 else "nonrevC.xlsx")
